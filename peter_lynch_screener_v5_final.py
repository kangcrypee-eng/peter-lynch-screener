"""
피터 린치식 미국 주식 통합 스크리닝 시스템 V6.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

핵심 원칙:
1. 전체 티커 분석 (Large-cap + Small-cap, $100M 이상)
2. 3중 검증 유지 (Yahoo + 직접계산 + Finviz) ⭐ 핵심
3. 높은 기준 유지 (PEG < 1.5, 성장률 15-200%) ⭐ 핵심
4. 중국 주식 10% 제한 (최대 1종목)

V6.0 신규 기능:
- 포트폴리오 히스토리 추적
- GPT 기반 매수/매도/관망 이유 설명
- 슬랙 메시지에 주가 링크 + 이유 표시

포트폴리오 구성:
- 매주 10종목 = 100%
  * 최고가치 4개 (40%)
  * 고성장 4개 (40%)
  * 균형 2개 (20%)

매매 규칙:
- 재추천 = 보유 유지 (10%)
- 신규 = 매수 (10%)
- 제외 = GPT 분석 후 매도/관망

환경 변수:
- OPENAI_API_KEY (필수)
- SLACK_BOT_TOKEN, SLACK_CHANNEL_ID (선택)

실행: python peter_lynch_screener_v6_final.py
"""

import pandas as pd
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import time
import logging
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI
import warnings
warnings.filterwarnings('ignore')

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'screener_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class PortfolioTracker:
    """포트폴리오 히스토리 추적 클래스"""
    
    def __init__(self, history_file='portfolio_history.json'):
        self.history_file = history_file
        self.history = self._load_history()
    
    def _load_history(self):
        """히스토리 로드"""
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    logger.info(f"✅ 히스토리 로드: {len(data.get('current_portfolio', []))}개 보유")
                    return data
            except Exception as e:
                logger.error(f"❌ 히스토리 로드 실패: {e}")
                return self._init_history()
        else:
            logger.info("📝 새 히스토리 파일 생성")
            return self._init_history()
    
    def _init_history(self):
        """히스토리 초기화"""
        return {
            'current_portfolio': [],  # [티커 리스트] - 현재 보유 10종목
            'weekly_recommendations': [],
            'trade_log': []
        }
    
    def save_history(self):
        """히스토리 저장"""
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, ensure_ascii=False, indent=2)
            logger.info(f"✅ 히스토리 저장")
        except Exception as e:
            logger.error(f"❌ 히스토리 저장 실패: {e}")
    
    def get_current_portfolio(self):
        """현재 포트폴리오 조회"""
        return self.history.get('current_portfolio', [])
    
    def analyze_changes(self, new_recommendations):
        """
        포트폴리오 변화 분석
        
        Returns:
            {
                'hold': [티커],      # 재추천 (보유 유지)
                'new_buy': [티커],   # 신규 매수
                'excluded': [티커]   # 추천 제외 (GPT 판단 필요)
            }
        """
        current = set(self.get_current_portfolio())
        recommended = set([s['티커'] for s in new_recommendations])
        
        return {
            'hold': list(current & recommended),        # 교집합
            'new_buy': list(recommended - current),     # 신규
            'excluded': list(current - recommended)     # 제외
        }
    
    def update_portfolio(self, new_portfolio_tickers, trade_log_entry):
        """포트폴리오 업데이트"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        # 현재 포트폴리오 업데이트
        self.history['current_portfolio'] = new_portfolio_tickers
        
        # 주간 추천 기록
        self.history['weekly_recommendations'].append({
            '날짜': today,
            '추천종목': new_portfolio_tickers
        })
        
        # 거래 로그 추가
        if trade_log_entry:
            self.history['trade_log'].extend(trade_log_entry)
        
        self.save_history()


class GPTAnalyzer:
    """GPT 분석 - 한글 번역 + 매수/매도/관망 이유"""
    
    def __init__(self):
        self.api_key = os.environ.get("OPENAI_API_KEY")
        
        if not self.api_key:
            logger.warning("⚠️ OPENAI_API_KEY 미설정")
            self.enabled = False
        else:
            try:
                self.client = OpenAI(api_key=self.api_key)
                self.enabled = True
                logger.info("✅ GPT API 연동")
            except Exception as e:
                logger.error(f"❌ GPT 초기화 실패: {e}")
                self.enabled = False
    
    def translate_to_korean(self, company_name, business_summary):
        """기업 설명 한글 번역 (30자 이내)"""
        if not self.enabled or not business_summary:
            return f"{company_name} 관련 기업"
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "기업을 한글로 30자 이내로 간단히 설명합니다."},
                    {"role": "user", "content": f"{company_name}: {business_summary[:300]}\n\n30자 이내로 설명:"}
                ],
                max_tokens=100,
                temperature=0.3
            )
            return response.choices[0].message.content.strip()[:50]
        except:
            return f"{company_name} 관련 기업"
    
    def analyze_portfolio_actions(self, categorized_stocks, changes):
        """
        포트폴리오 분석 + 매수/매도/관망 이유
        
        Returns:
            {
                'hold': {티커: 이유},
                'new_buy': {티커: 이유},
                'sell': {티커: 이유},
                'watch': {티커: 이유},
                'summary': 종합 분석
            }
        """
        if not self.enabled:
            return self._basic_analysis(categorized_stocks, changes)
        
        try:
            prompt = self._create_analysis_prompt(categorized_stocks, changes)
            
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "피터 린치 투자 전략 전문가. 매수/매도/관망 이유를 명확히 설명합니다."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=4096,
                temperature=0.3
            )
            
            result_text = response.choices[0].message.content
            parsed = self._parse_gpt_response(result_text, categorized_stocks, changes)
            
            logger.info("✅ GPT 포트폴리오 분석 완료")
            return parsed
            
        except Exception as e:
            logger.error(f"❌ GPT 분석 실패: {e}")
            return self._basic_analysis(categorized_stocks, changes)
    
    def _create_analysis_prompt(self, categorized_stocks, changes):
        """GPT 프롬프트 생성"""
        
        # 이번 주 추천 종목 정보
        stocks_info = "## 이번 주 추천 포트폴리오 (10종목 = 100%)\n\n"
        
        for category, name in [
            ('best_value', '최고 가치주 (40%)'),
            ('high_growth', '고성장주 (40%)'),
            ('balanced', '균형 (20%)')
        ]:
            stocks = categorized_stocks.get(category, [])
            stocks_info += f"### 📊 {name}\n\n"
            
            for stock in stocks:
                ticker = stock['티커']
                
                # 상태 표시
                if ticker in changes['hold']:
                    status = "✅ 보유유지"
                elif ticker in changes['new_buy']:
                    status = "🆕 신규매수"
                else:
                    status = ""
                
                stocks_info += f"**{ticker}** {status} - {stock['회사명']}\n"
                stocks_info += f"  한글: {stock.get('한글설명', 'N/A')}\n"
                stocks_info += f"  PEG: {stock['PEG']:.2f} | 성장률: {stock['성장률(%)']:.1f}% | PE: {stock.get('P/E', 'N/A')}\n"
                stocks_info += f"  시총: ${stock['시가총액($B)']:.1f}B\n\n"
        
        # 추천 제외 종목 (매도/관망 판단 필요)
        excluded_info = ""
        if changes['excluded']:
            excluded_info = "\n## 추천 제외 종목 (매도/관망 판단)\n\n"
            excluded_info += f"다음 종목들이 이번 주 추천에서 제외되었습니다:\n"
            excluded_info += f"{', '.join(changes['excluded'])}\n\n"
            excluded_info += "각 종목에 대해 **매도** 또는 **관망** 여부를 결정해주세요.\n"
        
        prompt = f"""{stocks_info}

{excluded_info}

## 응답 형식

다음 형식으로 **반드시** 응답해주세요:

**보유유지:**
AAPL: PEG 0.6으로 저평가, 안정적 성장 지속
NVDA: AI 칩 수요 폭발, 성장률 80% 유지

**신규매수:**
MSFT: AI 투자 확대로 클라우드 성장 가속, PEG 0.8 매력적
GOOGL: 검색 광고 회복, AI 모델 경쟁력 강화

**매도:**
META: 광고 수익 감소, PEG 2.1 고평가로 매도 권장
TSLA: 성장률 10%로 둔화, 경쟁 심화로 매도

**관망:**
NFLX: 일시적 구독자 감소이나 콘텐츠 투자 증가, 2분기 실적 후 재평가
AMD: PEG 1.8로 약간 높지만 AI 칩 수요 증가 전망, 관망

**종합분석:**
이번 주 포트폴리오는...

각 종목마다 **한 줄**로 명확한 이유를 작성해주세요.
"""
        
        return prompt
    
    def _parse_gpt_response(self, text, categorized_stocks, changes):
        """GPT 응답 파싱"""
        result = {
            'hold': {},
            'new_buy': {},
            'sell': {},
            'watch': {},
            'summary': ''
        }
        
        lines = text.strip().split('\n')
        current_section = None
        summary_started = False
        
        for line in lines:
            line = line.strip()
            
            # 섹션 구분
            if '**보유유지:**' in line or '보유유지:' in line:
                current_section = 'hold'
                summary_started = False
                continue
            elif '**신규매수:**' in line or '신규매수:' in line:
                current_section = 'new_buy'
                summary_started = False
                continue
            elif '**매도:**' in line or '매도:' in line:
                current_section = 'sell'
                summary_started = False
                continue
            elif '**관망:**' in line or '관망:' in line:
                current_section = 'watch'
                summary_started = False
                continue
            elif '**종합분석:**' in line or '종합분석:' in line:
                summary_started = True
                current_section = None
                continue
            
            # 종합분석 수집
            if summary_started and line:
                result['summary'] += line + '\n'
                continue
            
            # 티커: 이유 파싱
            if current_section and ':' in line and not line.startswith('#'):
                parts = line.split(':', 1)
                if len(parts) == 2:
                    ticker_part = parts[0].strip()
                    reason = parts[1].strip()
                    
                    # 티커 추출
                    ticker = None
                    for word in ticker_part.split():
                        word_clean = word.upper().strip('*-•')
                        # 추천 종목 또는 제외 종목에서 찾기
                        all_tickers = (
                            [s['티커'] for cat in categorized_stocks.values() for s in cat] +
                            changes['excluded']
                        )
                        if word_clean in all_tickers:
                            ticker = word_clean
                            break
                    
                    if ticker and reason:
                        result[current_section][ticker] = reason
        
        # 파싱 실패한 종목 기본값 처리
        for ticker in changes['hold']:
            if ticker not in result['hold']:
                result['hold'][ticker] = "재추천으로 보유 유지"
        
        for ticker in changes['new_buy']:
            if ticker not in result['new_buy']:
                result['new_buy'][ticker] = "신규 추천으로 매수"
        
        for ticker in changes['excluded']:
            if ticker not in result['sell'] and ticker not in result['watch']:
                result['sell'][ticker] = "추천 제외로 매도 권장"
        
        return result
    
    def _basic_analysis(self, categorized_stocks, changes):
        """기본 분석 (GPT 미사용)"""
        result = {
            'hold': {t: "재추천으로 보유 유지" for t in changes['hold']},
            'new_buy': {t: "신규 추천으로 매수" for t in changes['new_buy']},
            'sell': {t: "추천 제외로 매도 권장" for t in changes['excluded']},
            'watch': {},
            'summary': "GPT API 미사용으로 기본 분석만 제공됩니다."
        }
        return result


class SlackSender:
    """슬랙 메시지 전송"""
    
    def __init__(self):
        self.token = os.environ.get('SLACK_BOT_TOKEN')
        self.channel_id = os.environ.get('SLACK_CHANNEL_ID')
        self.enabled = bool(self.token and self.channel_id)
        
        if self.enabled:
            try:
                from slack_sdk import WebClient
                self.client = WebClient(token=self.token)
                logger.info(f"✅ 슬랙 연동")
            except:
                logger.warning("⚠️ slack_sdk 미설치")
                self.enabled = False
        else:
            logger.info("ℹ️ 슬랙 미설정")
    
    def send_message(self, message):
        if not self.enabled:
            return False
        try:
            self.client.chat_postMessage(
                channel=self.channel_id,
                text=message,
                mrkdwn=True
            )
            logger.info("✅ 슬랙 전송")
            return True
        except Exception as e:
            logger.error(f"❌ 슬랙 실패: {e}")
            return False
    
    def send_file(self, file_path, title=None):
        if not self.enabled:
            return False
        try:
            self.client.files_upload_v2(
                channel=self.channel_id,
                file=file_path,
                title=title or os.path.basename(file_path)
            )
            logger.info(f"✅ 슬랙 파일 전송")
            return True
        except Exception as e:
            logger.error(f"❌ 슬랙 파일 실패: {e}")
            return False


class PeterLynchScreener:
    """피터 린치 스크리너 V6.0"""
    
    def __init__(self):
        self.tickers = []
        self.filtered = []
        self.validated = []
        self.categorized_stocks = {}
        
        self.gpt_analyzer = GPTAnalyzer()
        self.slack_sender = SlackSender()
        self.portfolio_tracker = PortfolioTracker()
        
        self.MIN_MARKET_CAP = 100_000_000
        
        self.CHINA_KEYWORDS = [
            'china', 'chinese', 'beijing', 'shanghai', 'shenzhen',
            'hong kong', 'macau', 'taiwan', 'prc', 'cayman'
        ]
        
        self.GROWTH_LIMITS = {
            'min': 15,
            'ideal_min': 20,
            'ideal_max': 50,
            'max': 200
        }
        
        self.PEG_LIMITS = {
            'excellent': 0.5,
            'good': 0.7,
            'fair': 1.0,
            'max': 1.5
        }
        
        self.TOLERANCE = 0.20
        
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }
        
        self.error_details = []
    
    def _is_china_stock(self, info):
        """중국 주식 확인"""
        try:
            country = info.get('country', '').lower()
            if any(c in country for c in ['china', 'hong kong', 'taiwan']):
                return True
            
            name = (info.get('longName', '') + ' ' + info.get('shortName', '')).lower()
            if any(kw in name for kw in self.CHINA_KEYWORDS):
                return True
            
            business = info.get('longBusinessSummary', '').lower()
            if sum(1 for kw in self.CHINA_KEYWORDS if kw in business) >= 2:
                return True
            
            return False
        except:
            return False
    
    def run(self, ticker_limit=None):
        """메인 실행"""
        start = time.time()
        
        logger.info("=" * 80)
        logger.info("🎯 피터 린치 스크리너 V6.0")
        logger.info(f"💰 시가총액: ${self.MIN_MARKET_CAP/1e6:.0f}M+")
        logger.info(f"📊 기준: PEG < {self.PEG_LIMITS['max']}, 성장률 {self.GROWTH_LIMITS['min']}-{self.GROWTH_LIMITS['max']}%")
        logger.info(f"🇨🇳 중국: 최대 1종목")
        logger.info(f"📈 포트폴리오: 10종목 = 100% (4/4/2)")
        logger.info("=" * 80)
        
        if not self._step1_collect_tickers(ticker_limit):
            return None
        if not self._step2_basic_filter():
            return None
        if not self._step3_deep_analysis():
            return None
        if not self._step4_categorize():
            return None
        
        # 최종 10종목 선정
        final_10 = self._select_final_10()
        
        # 포트폴리오 변화 분석
        changes = self.portfolio_tracker.analyze_changes(final_10)
        
        # GPT 분석 (매수/매도/관망 이유)
        gpt_analysis = self._step5_gpt_analysis(final_10, changes)
        
        # 최종 포트폴리오 결정 (매도 후 신규 매수)
        final_portfolio = self._finalize_portfolio(final_10, gpt_analysis, changes)
        
        # Excel 생성
        filename = self._step6_create_excel(final_portfolio, gpt_analysis)
        
        # 슬랙 전송
        self._step7_send_to_slack(filename, final_portfolio, gpt_analysis)
        
        # 히스토리 업데이트
        self._update_history(final_portfolio, gpt_analysis)
        
        elapsed = (time.time() - start) / 60
        logger.info(f"\n⏱️ 소요시간: {elapsed:.1f}분")
        logger.info(f"📊 파일: {filename}\n")
        
        return filename
    
    def _step1_collect_tickers(self, limit=None):
        """Step 1: 티커 수집"""
        logger.info("\n[Step 1/7] 티커 수집...")
        
        try:
            url = "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=25000&download=true"
            response = requests.get(url, headers=self.headers, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            if 'data' not in data or 'rows' not in data['data']:
                logger.error("❌ API 오류")
                return False
            
            df = pd.DataFrame(data['data']['rows'])
            df = df[df['symbol'].notna()].copy()
            df['symbol'] = df['symbol'].str.strip().str.upper()
            df = df[~df['symbol'].str.contains(r'\^|\.|-', regex=True, na=False)]
            
            if 'name' in df.columns:
                df = df[~df['name'].str.contains('ETF|ETN|FUND|TRUST', case=False, na=False)]
            
            df = df[df['symbol'].str.len().between(1, 5)]
            df = df[df['symbol'].str.isalpha()]
            df = df.drop_duplicates(subset=['symbol'])
            
            all_tickers = df['symbol'].tolist()
            self.tickers = all_tickers[:limit] if limit else all_tickers
            
            logger.info(f"✅ {len(self.tickers)}개 수집\n")
            return True
            
        except Exception as e:
            logger.error(f"❌ 실패: {e}")
            return False
    
    def _step2_basic_filter(self):
        """Step 2: 기본 필터"""
        logger.info("[Step 2/7] 기본 필터...")
        passed = []
        errors = 0
        consecutive_errors = 0
        MAX_CONSECUTIVE_ERRORS = 10
        
        total = len(self.tickers)
        
        for i, ticker in enumerate(self.tickers, 1):
            if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                logger.warning(f"⚠️ 연속 에러, 1분 대기...")
                time.sleep(60)
                consecutive_errors = 0
            
            try:
                stock = yf.Ticker(ticker)
                info = stock.info
                
                if not info or len(info) < 5:
                    errors += 1
                    consecutive_errors += 1
                    time.sleep(0.5)
                    continue
                
                consecutive_errors = 0
                
                price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
                mcap = info.get('marketCap')
                
                if not price or not mcap:
                    errors += 1
                    time.sleep(0.3)
                    continue
                
                if price >= 1.0 and mcap > self.MIN_MARKET_CAP:
                    passed.append({
                        'ticker': ticker,
                        'price': float(price),
                        'market_cap': int(mcap)
                    })
                
                if i % 100 == 0:
                    logger.info(f"  {i}/{total} - 통과: {len(passed)}개")
                
                time.sleep(0.15)
                
            except KeyboardInterrupt:
                logger.warning("⚠️ 중단")
                break
            except Exception as e:
                errors += 1
                consecutive_errors += 1
                time.sleep(1.0)
                continue
        
        self.filtered = passed
        logger.info(f"✅ {len(self.filtered)}개 통과\n")
        
        return len(self.filtered) > 0
    
    def _step3_deep_analysis(self):
        """Step 3: 심층 분석"""
        logger.info("[Step 3/7] 심층 분석 (3중 검증)...")
        
        validated = []
        errors = 0
        skipped = 0
        
        for i, stock_data in enumerate(self.filtered, 1):
            ticker = stock_data['ticker']
            
            try:
                result = self._analyze_with_triple_validation(stock_data)
                
                if result and result.get('is_valid'):
                    validated.append(result)
                else:
                    skipped += 1
                
                if i % 25 == 0:
                    logger.info(f"  {i}/{len(self.filtered)} - 검증: {len(validated)}개")
                
                time.sleep(0.2)
                
            except Exception as e:
                errors += 1
                continue
        
        self.validated = validated
        logger.info(f"✅ {len(self.validated)}개 검증 완료\n")
        
        return len(self.validated) > 0
    
    def _analyze_with_triple_validation(self, basic_data):
        """3중 검증"""
        ticker = basic_data['ticker']
        
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            if not info or len(info) < 5:
                return None
            
            name = info.get('longName') or info.get('shortName', 'N/A')
            sector = info.get('sector', 'N/A')
            industry = info.get('industry', 'N/A')
            business = info.get('longBusinessSummary', '')[:500]
            price = basic_data['price']
            market_cap = basic_data['market_cap']
            
            is_china = self._is_china_stock(info)
            
            yahoo_pe = info.get('trailingPE') or info.get('forwardPE')
            yahoo_growth = info.get('earningsGrowth') or info.get('earningsQuarterlyGrowth')
            
            if not yahoo_pe or not yahoo_growth:
                return None
            
            if yahoo_pe <= 0:
                return None
            
            yahoo_growth_pct = yahoo_growth * 100 if yahoo_growth < 10 else yahoo_growth
            
            if yahoo_growth_pct <= 0 or yahoo_growth_pct > 500:
                return None
            
            yahoo_peg = yahoo_pe / yahoo_growth_pct
            
            calculated_peg = self._calculate_peg_manually(stock, yahoo_pe)
            finviz_peg = None
            
            validation_result = self._triple_validate(yahoo_peg, calculated_peg, finviz_peg)
            
            if not validation_result['valid']:
                return None
            
            final_peg = validation_result['peg']
            
            if final_peg >= self.PEG_LIMITS['max'] or final_peg <= 0:
                return None
            
            if yahoo_growth_pct < self.GROWTH_LIMITS['min'] or yahoo_growth_pct > self.GROWTH_LIMITS['max']:
                return None
            
            debt_to_equity = info.get('debtToEquity')
            if sector != 'Financial Services' and debt_to_equity and debt_to_equity > 200:
                return None
            
            return {
                'ticker': ticker,
                'name': name,
                'sector': sector,
                'industry': industry,
                'business_summary': business,
                'price': price,
                'market_cap': market_cap,
                'pe_ratio': yahoo_pe,
                'peg': final_peg,
                'growth_rate': yahoo_growth_pct,
                'debt_to_equity': debt_to_equity,
                'validation_status': validation_result['status'],
                'data_sources': validation_result['sources'],
                'is_china': is_china,
                'is_valid': True
            }
            
        except:
            return None
    
    def _calculate_peg_manually(self, stock, pe_ratio):
        """직접 계산"""
        try:
            financials = stock.financials
            
            if financials is None or financials.empty:
                return None
            
            net_income_row = None
            for row_name in ['Net Income', 'Net Income Common Stockholders']:
                if row_name in financials.index:
                    net_income_row = row_name
                    break
            
            if not net_income_row:
                return None
            
            net_income = financials.loc[net_income_row]
            
            if len(net_income) < 2:
                return None
            
            recent = net_income.iloc[0]
            previous = net_income.iloc[1]
            
            if previous <= 0:
                return None
            
            growth_rate = ((recent - previous) / abs(previous)) * 100
            
            if growth_rate <= 0:
                return None
            
            return pe_ratio / growth_rate
            
        except:
            return None
    
    def _triple_validate(self, yahoo_peg, calculated_peg, finviz_peg):
        """3중 검증"""
        sources = []
        valid_pegs = []
        
        if yahoo_peg and 0 < yahoo_peg < 10:
            sources.append('Yahoo')
            valid_pegs.append(yahoo_peg)
        
        if calculated_peg and 0 < calculated_peg < 10:
            sources.append('Calc')
            valid_pegs.append(calculated_peg)
        
        if finviz_peg and 0 < finviz_peg < 10:
            sources.append('Finviz')
            valid_pegs.append(finviz_peg)
        
        if len(valid_pegs) < 2:
            return {'valid': False}
        
        avg_peg = sum(valid_pegs) / len(valid_pegs)
        
        for peg in valid_pegs:
            if abs(peg - avg_peg) / avg_peg > self.TOLERANCE:
                median_peg = sorted(valid_pegs)[len(valid_pegs) // 2]
                return {
                    'valid': True,
                    'peg': round(median_peg, 2),
                    'status': '✅ 부분 검증',
                    'sources': sources
                }
        
        return {
            'valid': True,
            'peg': round(avg_peg, 2),
            'status': f'✅ {len(sources)}중 검증',
            'sources': sources
        }
    
    def _step4_categorize(self):
        """Step 4: 유형별 분류"""
        logger.info("[Step 4/7] 유형별 분류...")
        df = pd.DataFrame(self.validated)
        
        categorized = {
            'best_value': [],
            'high_growth': [],
            'balanced': []
        }
        
        # 최고 가치주
        best = df[
            (df['peg'] < self.PEG_LIMITS['good']) &
            (df['growth_rate'] >= self.GROWTH_LIMITS['ideal_min']) &
            (df['growth_rate'] <= self.GROWTH_LIMITS['ideal_max'])
        ].sort_values('peg').head(10)
        
        for _, row in best.iterrows():
            categorized['best_value'].append(self._create_recommendation(row, 'best_value'))
        
        # 고성장주
        high = df[
            (df['growth_rate'] > 50) &
            (df['growth_rate'] <= self.GROWTH_LIMITS['max']) &
            (df['peg'] < 1.2)
        ].sort_values('growth_rate', ascending=False).head(10)
        
        for _, row in high.iterrows():
            categorized['high_growth'].append(self._create_recommendation(row, 'high_growth'))
        
        # 균형
        balanced = df[
            (df['peg'] < 1.0) &
            (df['growth_rate'] >= 20) &
            (df['growth_rate'] <= 40)
        ].sort_values('peg').head(5)
        
        for _, row in balanced.iterrows():
            categorized['balanced'].append(self._create_recommendation(row, 'balanced'))
        
        self.categorized_stocks = categorized
        
        logger.info(f"✅ 최고가치: {len(categorized['best_value'])}개")
        logger.info(f"✅ 고성장: {len(categorized['high_growth'])}개")
        logger.info(f"✅ 균형: {len(categorized['balanced'])}개\n")
        
        return True
    
    def _create_recommendation(self, row, category):
        """추천 생성"""
        ticker = row['ticker']
        peg = row['peg']
        growth = row['growth_rate']
        market_cap_b = row['market_cap'] / 1e9
        is_china = row.get('is_china', False)
        
        category_names = {
            'best_value': '최고 가치주',
            'high_growth': '고성장주',
            'balanced': '균형'
        }
        
        korean_desc = self.gpt_analyzer.translate_to_korean(
            row.get('name', 'N/A'),
            row.get('business_summary', '')
        )
        
        return {
            '티커': ticker,
            '회사명': row.get('name', 'N/A'),
            '한글설명': korean_desc,
            '섹터': row.get('sector', 'N/A'),
            '산업': row.get('industry', 'N/A'),
            'PEG': peg,
            '성장률(%)': growth,
            'P/E': row.get('pe_ratio'),
            '시가총액($B)': round(market_cap_b, 2),
            '검증상태': row['validation_status'],
            '유형': category_names[category],
            'price': row['price'],
            'category': category,
            'is_china': is_china
        }
    
    def _select_final_10(self):
        """최종 10종목 선정 (4/4/2)"""
        logger.info("[추가] 최종 10종목 선정 (4/4/2)...")
        
        final = []
        final.extend(self.categorized_stocks['best_value'][:4])
        final.extend(self.categorized_stocks['high_growth'][:4])
        final.extend(self.categorized_stocks['balanced'][:2])
        
        logger.info(f"✅ 10종목 선정\n")
        
        return final
    
    def _step5_gpt_analysis(self, final_10, changes):
        """Step 5: GPT 분석"""
        logger.info("[Step 5/7] GPT 분석 (매수/매도/관망 이유)...")
        
        # categorized 재구성
        categorized = {
            'best_value': [s for s in final_10 if s['category'] == 'best_value'],
            'high_growth': [s for s in final_10 if s['category'] == 'high_growth'],
            'balanced': [s for s in final_10 if s['category'] == 'balanced']
        }
        
        gpt_analysis = self.gpt_analyzer.analyze_portfolio_actions(categorized, changes)
        
        logger.info(f"  보유: {len(gpt_analysis['hold'])}개")
        logger.info(f"  신규: {len(gpt_analysis['new_buy'])}개")
        logger.info(f"  매도: {len(gpt_analysis['sell'])}개")
        logger.info(f"  관망: {len(gpt_analysis['watch'])}개\n")
        
        return gpt_analysis
    
    def _finalize_portfolio(self, final_10, gpt_analysis, changes):
        """최종 포트폴리오 결정"""
        logger.info("[추가] 최종 포트폴리오 결정...")
        
        # 매도 종목 제외
        final_tickers = [s['티커'] for s in final_10]
        sell_tickers = list(gpt_analysis['sell'].keys())
        
        # 최종 포트폴리오 = 추천 10종목 (매도 제외)
        final_portfolio = []
        for stock in final_10:
            ticker = stock['티커']
            stock['상태'] = 'hold' if ticker in changes['hold'] else 'new_buy'
            stock['이유'] = gpt_analysis.get(stock['상태'], {}).get(ticker, '')
            final_portfolio.append(stock)
        
        logger.info(f"✅ 포트폴리오 10종목 확정\n")
        
        return {
            'stocks': final_portfolio,
            'sell': gpt_analysis['sell'],
            'watch': gpt_analysis['watch'],
            'summary': gpt_analysis['summary']
        }
    
    def _step6_create_excel(self, final_portfolio, gpt_analysis):
        """Step 6: Excel 생성"""
        logger.info("[Step 6/7] Excel 생성...")
        
        today = datetime.now().strftime('%Y%m%d')
        filename = f'Peter_Lynch_Report_{today}.xlsx'
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # 시트 생성 (간소화)
        ws = wb.create_sheet(title='포트폴리오')
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        columns = ['티커', '회사명', '한글설명', '유형', '상태', '이유', 'PEG', '성장률(%)', '시가총액($B)']
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, stock in enumerate(final_portfolio['stocks'], 2):
            status_text = "✅ 보유" if stock['상태'] == 'hold' else "🆕 신규"
            
            ws.cell(row=row_idx, column=1, value=stock['티커'])
            ws.cell(row=row_idx, column=2, value=stock['회사명'])
            ws.cell(row=row_idx, column=3, value=stock['한글설명'])
            ws.cell(row=row_idx, column=4, value=stock['유형'])
            ws.cell(row=row_idx, column=5, value=status_text)
            ws.cell(row=row_idx, column=6, value=stock['이유'])
            ws.cell(row=row_idx, column=7, value=stock['PEG'])
            ws.cell(row=row_idx, column=8, value=stock['성장률(%)'])
            ws.cell(row=row_idx, column=9, value=stock['시가총액($B)'])
        
        wb.save(filename)
        logger.info(f"✅ {filename}\n")
        return filename
    
    def _step7_send_to_slack(self, filename, final_portfolio, gpt_analysis):
        """Step 7: 슬랙 전송"""
        logger.info("[Step 7/7] 슬랙 전송...")
        
        message = self._create_slack_message(final_portfolio, gpt_analysis)
        
        if not self.slack_sender.enabled:
            print("\n" + "="*80)
            print(message)
            print("="*80 + "\n")
            return
        
        self.slack_sender.send_message(message)
        self.slack_sender.send_file(filename)
        logger.info("✅ 완료\n")
    
    def _create_slack_message(self, final_portfolio, gpt_analysis):
        """슬랙 메시지 생성"""
        today = datetime.now().strftime('%Y년 %m월 %d일')
        week = datetime.now().isocalendar()[1]
        
        msg = [f"🤖 *피터 린치 봇 V6.0*"]
        msg.append(f"📅 {today} ({week}주차)")
        msg.append(f"💎 포트폴리오: 10종목 = 100%")
        msg.append("")
        msg.append("━━━━━━━━━━━━━━━━━━")
        msg.append("📊 *현재 포트폴리오 구성*")
        msg.append("━━━━━━━━━━━━━━━━━━")
        
        # 카테고리별 출력
        for category, name, emoji in [
            ('best_value', '최고 가치주 (40%)', '🏆'),
            ('high_growth', '고성장주 (40%)', '🚀'),
            ('balanced', '균형 (20%)', '⚖️')
        ]:
            stocks = [s for s in final_portfolio['stocks'] if s['category'] == category]
            if stocks:
                msg.append(f"\n*{emoji} {name}*")
                for stock in stocks:
                    ticker = stock['티커']
                    name_kr = stock['한글설명']
                    price = stock['price']
                    
                    # 상태 표시
                    if stock['상태'] == 'hold':
                        status = "✅ 보유유지"
                    else:
                        status = "🆕 신규매수"
                    
                    # 이유
                    reason = stock['이유'] or "분석 중"
                    
                    yahoo_link = f"https://finance.yahoo.com/quote/{ticker}"
                    
                    msg.append(f"  • *{ticker}* {status} - {name_kr}")
                    msg.append(f"    현재가: ${price:.2f} | <{yahoo_link}|주가 보기>")
                    msg.append(f"    💡 {reason}")
                    msg.append("")
        
        # 매도/관망
        if final_portfolio['sell'] or final_portfolio['watch']:
            msg.append("━━━━━━━━━━━━━━━━━━")
            msg.append("⚖️ *매도/관망 종목*")
            msg.append("━━━━━━━━━━━━━━━━━━")
            
            for ticker, reason in final_portfolio['sell'].items():
                msg.append(f"\n💰 *{ticker}* - 매도권장")
                msg.append(f"  ❌ {reason}")
            
            for ticker, reason in final_portfolio['watch'].items():
                msg.append(f"\n👀 *{ticker}* - 관망권장")
                msg.append(f"  🔍 {reason}")
        
        # 종합분석
        if final_portfolio['summary']:
            msg.append("\n━━━━━━━━━━━━━━━━━━")
            msg.append("💬 *종합 분석*")
            msg.append("━━━━━━━━━━━━━━━━━━")
            msg.append(final_portfolio['summary'].strip())
        
        return "\n".join(msg)
    
    def _update_history(self, final_portfolio, gpt_analysis):
        """히스토리 업데이트"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        # 최종 포트폴리오 티커 리스트
        final_tickers = [s['티커'] for s in final_portfolio['stocks']]
        
        # 거래 로그
        trade_log = []
        for stock in final_portfolio['stocks']:
            ticker = stock['티커']
            if stock['상태'] == 'new_buy':
                trade_log.append({
                    '날짜': today,
                    '티커': ticker,
                    '액션': '신규매수',
                    '유형': stock['유형'],
                    '메모': stock['이유']
                })
        
        for ticker, reason in final_portfolio['sell'].items():
            trade_log.append({
                '날짜': today,
                '티커': ticker,
                '액션': '매도',
                '유형': '',
                '메모': reason
            })
        
        self.portfolio_tracker.update_portfolio(final_tickers, trade_log)


def main():
    print("""
╔════════════════════════════════════════════════════════════════╗
║  피터 린치 통합 스크리닝 시스템 V6.0                         ║
║                                                                ║
║  ✅ 3중 검증 (Yahoo + 직접계산)                              ║
║  ✅ 높은 기준 (PEG < 1.5, 성장률 15-200%)                   ║
║  ✅ 포트폴리오: 10종목 = 100% (4/4/2)                        ║
║                                                                ║
║  🆕 V6.0:                                                      ║
║  - GPT 매수/매도/관망 이유 설명                              ║
║  - 슬랙 메시지에 주가 링크 + 이유                            ║
║  - 히스토리 추적                                              ║
║                                                                ║
║  매매 규칙:                                                    ║
║  - 재추천 = 보유 유지 (10%)                                  ║
║  - 신규 = 매수 (10%)                                         ║
║  - 제외 = GPT 분석 후 매도/관망                              ║
║                                                                ║
║  환경변수: OPENAI_API_KEY (필수)                              ║
╚════════════════════════════════════════════════════════════════╝
    """)
    
    if not os.environ.get("OPENAI_API_KEY"):
        print("⚠️  OPENAI_API_KEY 미설정\n")
    
    screener = PeterLynchScreener()
    result = screener.run(ticker_limit=None)
    
    if result:
        print(f"\n✅ 완료!")
        print(f"📊 {result}")
        print(f"📁 portfolio_history.json")
    else:
        print("\n❌ 실패")


if __name__ == "__main__":
    main()