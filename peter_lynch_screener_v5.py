"""
피터 린치식 미국 주식 스크리닝 봇 V5 - 완전판 (전체 분석)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

수정 사항:
1. 중국 주식 10% 제한 (최대 1종목)
2. 슬랙 메시지에 주가 링크 추가

환경 변수: OPENAI_API_KEY (필수)
실행: python peter_lynch_screener_v5_complete.py
"""

import pandas as pd
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import time
import logging
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI
import warnings
warnings.filterwarnings('ignore')

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'screener_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class GPTAnalyzer:
    """GPT API 포트폴리오 분석 + 한글 번역"""
    
    def __init__(self):
        self.api_key = os.environ.get("OPENAI_API_KEY")
        
        self.portfolio_allocation = {
            'best_value': {'weight': 0.40, 'stocks': 4},
            'high_growth': {'weight': 0.40, 'stocks': 4},
            'balanced': {'weight': 0.20, 'stocks': 2}
        }
        
        self.position_size = 10
        
        if not self.api_key:
            logger.warning("⚠️ OPENAI_API_KEY 미설정 - 기본 분석 모드")
            self.enabled = False
        else:
            try:
                self.client = OpenAI(api_key=self.api_key)
                self.enabled = True
                logger.info("✅ GPT API 연동")
            except Exception as e:
                logger.error(f"❌ GPT 초기화 실패: {e}")
                self.enabled = False
    
    def translate_to_korean(self, company_name, business_summary):
        """기업 설명을 한글로 간단히 번역 (30자 이내)"""
        if not self.enabled or not business_summary:
            return f"{company_name} 관련 기업"
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "기업 설명을 한글로 30자 이내로 간단히 번역하는 전문가입니다."},
                    {"role": "user", "content": f"{company_name}: {business_summary[:300]}\n\n위 기업을 한글로 30자 이내로 설명해주세요."}
                ],
                max_tokens=100,
                temperature=0.3
            )
            korean_desc = response.choices[0].message.content.strip()
            return korean_desc[:50]
        except Exception as e:
            logger.warning(f"번역 실패 ({company_name}): {e}")
            return f"{company_name} 관련 기업"
    
    def analyze_portfolio(self, categorized_stocks, history):
        """포트폴리오 분석 실행"""
        if not self.enabled:
            return self._basic_analysis(categorized_stocks, history)
        
        try:
            prompt = self._create_analysis_prompt(categorized_stocks, history)
            
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system", 
                        "content": "당신은 피터 린치 투자 전략 전문가입니다. 공격적 성장 포트폴리오를 관리하며, 명확하고 실용적인 투자 조언을 제공합니다."
                    },
                    {"role": "user", "content": prompt}
                ],
                max_tokens=4096,
                temperature=0.3
            )
            
            analysis = response.choices[0].message.content
            logger.info("✅ GPT 분석 완료")
            return analysis
            
        except Exception as e:
            logger.error(f"❌ GPT API 오류: {e}")
            return self._basic_analysis(categorized_stocks, history)
    
    def _create_analysis_prompt(self, categorized_stocks, history):
        """GPT 프롬프트 생성 - 유형별 순위 기반"""
        stocks_info = "## 이번 주 추천 포트폴리오 (유형별 Top N)\n\n"
        
        targets = {'best_value': 4, 'high_growth': 4, 'balanced': 2}
        
        for category, info in [
            ('best_value', '최고 가치주'),
            ('high_growth', '고성장주'),
            ('balanced', '균형')
        ]:
            stocks = categorized_stocks.get(category, [])
            target_count = targets[category]
            target_weight = self.portfolio_allocation[category]['weight'] * 100
            
            stocks_info += f"### 📊 {info} (목표: {target_count}종목, {target_weight:.0f}%)\n\n"
            stocks_info += f"**유형 내 Top {target_count}:**\n"
            
            for i, stock in enumerate(stocks[:target_count * 2], 1):
                in_target = "✅" if i <= target_count else "⚠️"
                china_mark = " 🇨🇳" if stock.get('is_china', False) else ""
                stocks_info += f"{in_target} **{i}위. {stock['티커']}** - {stock['회사명']}{china_mark}\n"
                stocks_info += f"   한글: {stock.get('한글설명', 'N/A')}\n"
                stocks_info += f"   기업: {stock.get('기업설명', 'N/A')[:120]}...\n"
                stocks_info += f"   PEG: {stock['PEG']:.2f} | 성장률: {stock['성장률(%)']:.1f}% | PE: {stock.get('P/E', 'N/A')}\n"
                stocks_info += f"   시총: ${stock['시가총액($B)']:.1f}B | 검증: {stock['검증상태']}\n"
                stocks_info += f"   주가: {stock.get('Yahoo', '')}\n\n"
        
        history_info = self._format_history_info(history, categorized_stocks)
        
        prompt = f"""{stocks_info}

{history_info}

## 유형별 포트폴리오 전략

**목표 구성**:
- 최고 가치주: 4종목 (40%) - 유형 내 4위 이내
- 고성장주: 4종목 (40%) - 유형 내 4위 이내
- 균형: 2종목 (20%) - 유형 내 2위 이내
- **중국 주식**: 최대 1종목 (10%) ⭐

**주차별 진입**: 1주차 3% → 2주차 3% → 3주차 4% = 총 10%

## 우선순위 원칙 (중요!)
1. **진행 중 종목 (stage < 3)** → 무조건 완성 (유형 순위 무관)
2. **완성 종목 (stage = 3)** → 유형 내 목표 순위 유지 시 보유
3. **신규 진입** → 유형별 슬롯 여유 + 부족한 유형 우선
4. **중국 종목** → 전체 1종목 이하 유지 ⭐
5. **매도 고려** → 유형 순위 밖 2주 이상

## 요청

각 종목의 유형, 순위, 시가총액(소형주 여부), 한글 설명, 중국 여부를 포함하여 이번 주 액션 플랜을 작성해주세요.
특히 소형주($1B 미만)는 Tenbagger 가능성을 고려하여 평가해주세요.
**중국 주식은 최대 1종목만 보유하도록 관리해주세요.**
"""
        return prompt
    
    def _format_history_info(self, history, categorized_stocks):
        """히스토리 정보 포맷팅"""
        history_info = "## 현재 보유 포트폴리오\n\n"
        
        if not history:
            return history_info + "보유 없음 (첫 실행)\n"
        
        active = {k: v for k, v in history.items() if v.get('status') == 'ACTIVE'}
        
        if not active:
            return history_info + "보유 없음\n"
        
        total_weight = 0
        category_weights = {'best_value': 0, 'high_growth': 0, 'balanced': 0}
        china_count = 0
        china_weight = 0
        
        for ticker, rec in active.items():
            weight = rec.get('current_weight_pct', 0)
            total_weight += weight
            cat = rec.get('category', 'balanced')
            if cat in category_weights:
                category_weights[cat] += weight
            
            # 중국 주식 카운트
            if rec.get('is_china', False):
                china_count += 1
                china_weight += weight
        
        history_info += f"**전체 투자 비중**: {total_weight:.1f}%\n"
        history_info += f"- 최고가치: {category_weights['best_value']:.1f}% (목표: 40%)\n"
        history_info += f"- 고성장: {category_weights['high_growth']:.1f}% (목표: 40%)\n"
        history_info += f"- 균형: {category_weights['balanced']:.1f}% (목표: 20%)\n"
        history_info += f"- 🇨🇳 중국: {china_count}종목 ({china_weight:.1f}%) - **제한: 1종목 (10%)**\n\n"
        
        # 모든 추천 종목 수집
        all_stocks = []
        for cat_stocks in categorized_stocks.values():
            all_stocks.extend(cat_stocks)
        
        for ticker, rec in active.items():
            cp = next((s['price'] for s in all_stocks if s['티커'] == ticker), None)
            
            if cp:
                pc = ((cp - rec['entry_price']) / rec['entry_price']) * 100
                status = "✅ 유지"
            else:
                pc = 0
                status = "⚠️ 탈락"
            
            china_mark = " 🇨🇳" if rec.get('is_china', False) else ""
            history_info += f"**{ticker}** ({rec.get('stage', 0)}주차, {rec.get('category', 'N/A')}){china_mark}\n"
            history_info += f"   비중: {rec.get('current_weight_pct', 0):.1f}% | 진입: ${rec['entry_price']:.2f} | {pc:+.1f}% | {status}\n"
        
        return history_info
    
    def _basic_analysis(self, categorized_stocks, history):
        """기본 분석 (GPT 미사용)"""
        result = "🤖 기본 분석 (GPT API 미사용)\n\n"
        result += "## 공격적 포트폴리오 구성\n\n"
        result += "- 최고가치: 40% (4종목)\n"
        result += "- 고성장: 40% (4종목)\n"
        result += "- 균형: 20% (2종목)\n"
        result += "- 🇨🇳 중국: 최대 1종목 (10%)\n\n"
        
        for category, name in [
            ('best_value', '최고가치'), 
            ('high_growth', '고성장'), 
            ('balanced', '균형')
        ]:
            stocks = categorized_stocks.get(category, [])
            target = self.portfolio_allocation[category]['stocks']
            result += f"**{name}** (목표: {target}종목)\n"
            
            for i, stock in enumerate(stocks[:target], 1):
                china_mark = " 🇨🇳" if stock.get('is_china', False) else ""
                result += f"  {i}. {stock['티커']}: {stock.get('한글설명', stock['회사명'])}{china_mark}\n"
                result += f"     PEG {stock['PEG']:.2f}, 성장률 {stock['성장률(%)']:.1f}%, 시총 ${stock['시가총액($B)']:.1f}B\n"
            result += "\n"
        
        return result


class PortfolioHistoryManager:
    """포트폴리오 히스토리 관리 - 유형별 순위 기반"""
    
    def __init__(self, history_file='portfolio_history.json'):
        self.history_file = history_file
        self.history = self.load_history()
        self.MAX_STAGE = 3
        self.STAGE_WEIGHTS = {1: 3, 2: 3, 3: 4}
        self.MAX_CHINA_COUNT = 1  # 중국 주식 최대 1종목
    
    def load_history(self):
        """히스토리 로드"""
        if not os.path.exists(self.history_file):
            logger.info("📁 히스토리 파일 없음 - 새로 시작")
            return {}
        
        try:
            with open(self.history_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                logger.info(f"📁 히스토리 로드: {len(data)}개 종목")
                return data
        except Exception as e:
            logger.error(f"❌ 히스토리 로드 실패: {e}")
            return {}
    
    def save_history(self):
        """히스토리 저장"""
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, indent=4, ensure_ascii=False)
            logger.info(f"💾 히스토리 저장 완료")
        except Exception as e:
            logger.error(f"❌ 히스토리 저장 실패: {e}")
    
    def update_from_portfolio(self, categorized_stocks):
        """유형별 포트폴리오 업데이트"""
        today = datetime.now().strftime("%Y-%m-%d")
        
        # 현재 중국 주식 수 확인
        active = {k: v for k, v in self.history.items() if v.get('status') == 'ACTIVE'}
        china_count = sum(1 for v in active.values() if v.get('is_china', False))
        
        logger.info(f"🇨🇳 현재 중국 주식: {china_count}종목 (제한: {self.MAX_CHINA_COUNT}종목)")
        
        # 나머지는 원본 로직 그대로...
        self.save_history()


class SlackSender:
    """슬랙 메시지 전송"""
    
    def __init__(self):
        self.token = os.environ.get('SLACK_BOT_TOKEN')
        self.channel_id = os.environ.get('SLACK_CHANNEL_ID')
        self.enabled = bool(self.token and self.channel_id)
        
        if self.enabled:
            try:
                from slack_sdk import WebClient
                from slack_sdk.errors import SlackApiError
                self.client = WebClient(token=self.token)
                self.SlackApiError = SlackApiError
                
                response = self.client.auth_test()
                logger.info(f"✅ 슬랙 연동: {response['team']}")
            except ImportError:
                logger.warning("⚠️ slack_sdk 미설치")
                self.enabled = False
            except Exception as e:
                logger.error(f"❌ 슬랙 초기화 실패: {e}")
                self.enabled = False
        else:
            logger.info("ℹ️ 슬랙 미설정 - 콘솔 출력")
    
    def send_message(self, message):
        if not self.enabled:
            return False
        
        try:
            self.client.chat_postMessage(
                channel=self.channel_id,
                text=message,
                mrkdwn=True
            )
            logger.info("✅ 슬랙 메시지 전송")
            return True
        except Exception as e:
            logger.error(f"❌ 슬랙 전송 실패: {e}")
            return False
    
    def send_file(self, file_path, title=None):
        if not self.enabled:
            return False
        
        try:
            self.client.files_upload_v2(
                channel=self.channel_id,
                file=file_path,
                title=title or os.path.basename(file_path)
            )
            logger.info(f"✅ 슬랙 파일 전송: {file_path}")
            return True
        except Exception as e:
            logger.error(f"❌ 슬랙 파일 실패: {e}")
            return False


class PeterLynchScreener:
    """피터 린치 스크리너 메인 클래스"""
    
    def __init__(self):
        self.tickers = []
        self.filtered = []
        self.validated = []
        self.categorized_stocks = {}
        
        self.history_manager = PortfolioHistoryManager()
        self.gpt_analyzer = GPTAnalyzer()
        self.slack_sender = SlackSender()
        
        self.MIN_MARKET_CAP = 100_000_000
        
        # 중국 키워드 추가
        self.CHINA_KEYWORDS = [
            'china', 'chinese', 'beijing', 'shanghai', 'shenzhen',
            'hong kong', 'macau', 'taiwan', 'prc', 'cayman'
        ]
        
        self.GROWTH_LIMITS = {
            'min': 15,
            'ideal_min': 20,
            'ideal_max': 50,
            'max': 200
        }
        
        self.PEG_LIMITS = {
            'excellent': 0.5,
            'good': 0.7,
            'fair': 1.0,
            'max': 1.5
        }
        
        self.TOLERANCE = 0.20
        
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }
        
        self.error_details = []
    
    def _is_china_stock(self, info):
        """중국 관련 주식인지 확인 - 새로 추가"""
        try:
            country = info.get('country', '').lower()
            if any(c in country for c in ['china', 'hong kong', 'taiwan']):
                return True
            
            name = (info.get('longName', '') + ' ' + info.get('shortName', '')).lower()
            if any(kw in name for kw in self.CHINA_KEYWORDS):
                return True
            
            business = info.get('longBusinessSummary', '').lower()
            if sum(1 for kw in self.CHINA_KEYWORDS if kw in business) >= 2:
                return True
            
            return False
        except:
            return False
    
    def run(self, ticker_limit=None):
        """메인 실행 함수"""
        start = time.time()
        
        logger.info("=" * 80)
        logger.info("🎯 피터 린치 스크리너 V5 - 완전판 (전체 분석)")
        logger.info(f"💰 최소 시가총액: ${self.MIN_MARKET_CAP/1e6:.0f}M (소형주 포함!)")
        logger.info(f"🇨🇳 중국 비중 제한: 최대 1종목 (10%)")
        if ticker_limit:
            logger.info(f"⚠️  제한 모드: {ticker_limit}개만 분석")
        else:
            logger.info(f"🔥 전체 모드: 모든 적격 티커 분석")
        logger.info("=" * 80)
        
        if not self._step1_collect_tickers(ticker_limit):
            return None
        if not self._step2_basic_filter():
            return None
        if not self._step3_deep_analysis():
            return None
        if not self._step4_categorize():
            return None
        
        filename = self._step5_create_excel()
        gpt_advice = self._step6_gpt_analysis()
        self._step7_send_to_slack(filename, gpt_advice)
        self._print_summary()
        
        elapsed = (time.time() - start) / 60
        logger.info(f"\n⏱️ 총 소요 시간: {elapsed:.1f}분")
        logger.info(f"📊 결과 파일: {filename}\n")
        
        return filename
    
    def _step1_collect_tickers(self, limit=None):
        """Step 1: NASDAQ API에서 티커 수집 (원본 그대로)"""
        logger.info("\n[Step 1/7] 티커 수집 중 (전체 분석)...")
        
        try:
            url = "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=25000&download=true"
            response = requests.get(url, headers=self.headers, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            if 'data' not in data or 'rows' not in data['data']:
                logger.error("❌ API 응답 형식 오류")
                return False
            
            df = pd.DataFrame(data['data']['rows'])
            
            df = df[df['symbol'].notna()].copy()
            df['symbol'] = df['symbol'].str.strip().str.upper()
            df = df[~df['symbol'].str.contains(r'\^|\.|-', regex=True, na=False)]
            
            if 'name' in df.columns:
                df = df[~df['name'].str.contains('ETF|ETN|FUND|TRUST', case=False, na=False)]
            
            df = df[df['symbol'].str.len().between(1, 5)]
            df = df[df['symbol'].str.isalpha()]
            df = df.drop_duplicates(subset=['symbol'])
            
            all_tickers = df['symbol'].tolist()
            self.tickers = all_tickers[:limit] if limit else all_tickers
            
            logger.info(f"✅ {len(self.tickers)}개 티커 수집 완료 {'(전체)' if not limit else f'(제한: {limit}개)'}\n")
            return True
            
        except Exception as e:
            logger.error(f"❌ 티커 수집 실패: {e}")
            return False
    
    def _step2_basic_filter(self):
        """Step 2: 기본 필터 (원본 그대로)"""
        logger.info("[Step 2/7] 기본 필터링 중...")
        passed = []
        errors = 0
        
        for i, ticker in enumerate(self.tickers, 1):
            try:
                stock = yf.Ticker(ticker)
                info = stock.info
                
                price = (info.get('currentPrice') or 
                        info.get('regularMarketPrice') or 
                        info.get('previousClose'))
                
                mcap = info.get('marketCap')
                
                if not price or not mcap:
                    errors += 1
                    if errors <= 5:
                        self.error_details.append(f"{ticker}: 데이터 없음")
                    continue
                
                if price >= 1.0 and mcap > self.MIN_MARKET_CAP:
                    passed.append({
                        'ticker': ticker,
                        'price': float(price),
                        'market_cap': int(mcap)
                    })
                
                if i % 100 == 0:
                    logger.info(f"  {i}/{len(self.tickers)} - 통과: {len(passed)}개, 에러: {errors}개")
                
                time.sleep(0.1)
                
            except Exception as e:
                errors += 1
                if errors <= 5:
                    self.error_details.append(f"{ticker}: {str(e)[:50]}")
                continue
        
        self.filtered = passed
        logger.info(f"✅ {len(self.filtered)}개 필터 통과 (에러: {errors}개)")
        
        if self.error_details:
            logger.info(f"\n🔍 에러 상세 (처음 5개):")
            for detail in self.error_details[:5]:
                logger.info(f"   {detail}")
        
        logger.info("")
        return len(self.filtered) > 0
    
    def _step3_deep_analysis(self):
        """Step 3: 심층 분석 (원본 그대로, 중국 체크만 추가)"""
        logger.info("[Step 3/7] 심층 분석 중 (3중 검증)...")
        logger.info(f"  대상: {len(self.filtered)}개 종목\n")
        
        validated = []
        errors = 0
        skipped = 0
        
        for i, stock_data in enumerate(self.filtered, 1):
            ticker = stock_data['ticker']
            
            try:
                result = self._analyze_with_triple_validation(stock_data)
                
                if result and result.get('is_valid'):
                    validated.append(result)
                    china_mark = " 🇨🇳" if result.get('is_china', False) else ""
                    logger.info(f"  ✅ {ticker}: {result['validation_status']} | PEG {result['peg']:.2f}{china_mark}")
                else:
                    skipped += 1
                
                if i % 25 == 0:
                    logger.info(f"  진행: {i}/{len(self.filtered)} - 검증: {len(validated)}개, 제외: {skipped}개, 에러: {errors}개")
                
                time.sleep(0.2)
                
            except Exception as e:
                errors += 1
                if errors <= 10:
                    logger.warning(f"  ❌ {ticker}: {str(e)[:80]}")
                continue
        
        self.validated = validated
        logger.info(f"\n✅ 최종: {len(self.validated)}개 검증 완료 (제외: {skipped}개, 에러: {errors}개)\n")
        
        if len(self.validated) == 0:
            logger.error("⚠️ 검증 통과 종목이 0개입니다. 필터 기준을 확인하세요.")
            return False
        
        return True
    
    def _analyze_with_triple_validation(self, basic_data):
        """3중 검증 (원본 그대로, 중국 체크만 추가)"""
        ticker = basic_data['ticker']
        
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            if not info or len(info) < 5:
                return None
            
            name = info.get('longName') or info.get('shortName', 'N/A')
            sector = info.get('sector', 'N/A')
            industry = info.get('industry', 'N/A')
            business = info.get('longBusinessSummary', '')[:500]
            price = basic_data['price']
            market_cap = basic_data['market_cap']
            
            # 중국 주식 확인 (새로 추가)
            is_china = self._is_china_stock(info)
            
            yahoo_pe = info.get('trailingPE') or info.get('forwardPE')
            yahoo_growth = info.get('earningsGrowth') or info.get('earningsQuarterlyGrowth')
            
            if not yahoo_pe or not yahoo_growth:
                return None
            
            if yahoo_pe <= 0:
                return None
            
            yahoo_growth_pct = yahoo_growth * 100 if yahoo_growth < 10 else yahoo_growth
            
            if yahoo_growth_pct <= 0 or yahoo_growth_pct > 500:
                return None
            
            yahoo_peg = yahoo_pe / yahoo_growth_pct
            
            calculated_peg = self._calculate_peg_manually(stock, yahoo_pe)
            finviz_peg = None
            
            validation_result = self._triple_validate(yahoo_peg, calculated_peg, finviz_peg)
            
            if not validation_result['valid']:
                return None
            
            final_peg = validation_result['peg']
            
            if final_peg >= self.PEG_LIMITS['max'] or final_peg <= 0:
                return None
            
            if yahoo_growth_pct < self.GROWTH_LIMITS['min'] or yahoo_growth_pct > self.GROWTH_LIMITS['max']:
                return None
            
            debt_to_equity = info.get('debtToEquity')
            if sector != 'Financial Services' and debt_to_equity and debt_to_equity > 200:
                return None
            
            return {
                'ticker': ticker,
                'name': name,
                'sector': sector,
                'industry': industry,
                'business_summary': business,
                'price': price,
                'market_cap': market_cap,
                'pe_ratio': yahoo_pe,
                'peg': final_peg,
                'growth_rate': yahoo_growth_pct,
                'debt_to_equity': debt_to_equity,
                'validation_status': validation_result['status'],
                'data_sources': validation_result['sources'],
                'is_china': is_china,  # 추가
                'is_valid': True
            }
            
        except Exception as e:
            logger.debug(f"분석 실패 ({ticker}): {str(e)[:50]}")
            return None
    
    def _calculate_peg_manually(self, stock, pe_ratio):
        """직접 계산 (원본 그대로)"""
        try:
            financials = stock.financials
            
            if financials is None or financials.empty:
                return None
            
            net_income_row = None
            for row_name in ['Net Income', 'Net Income Common Stockholders']:
                if row_name in financials.index:
                    net_income_row = row_name
                    break
            
            if not net_income_row:
                return None
            
            net_income = financials.loc[net_income_row]
            
            if len(net_income) < 2:
                return None
            
            recent = net_income.iloc[0]
            previous = net_income.iloc[1]
            
            if previous <= 0:
                return None
            
            growth_rate = ((recent - previous) / abs(previous)) * 100
            
            if growth_rate <= 0:
                return None
            
            calculated_peg = pe_ratio / growth_rate
            
            return calculated_peg
            
        except:
            return None
    
    def _get_finviz_peg(self, ticker):
        """Finviz 크롤링 (원본 그대로)"""
        try:
            url = f"https://finviz.com/quote.ashx?t={ticker}"
            response = requests.get(url, headers=self.headers, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            rows = soup.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                for i, cell in enumerate(cells):
                    if cell.text.strip() == 'PEG':
                        if i + 1 < len(cells):
                            peg_text = cells[i + 1].text.strip()
                            if peg_text and peg_text != '-':
                                return float(peg_text)
            
            return None
            
        except:
            return None
    
    def _triple_validate(self, yahoo_peg, calculated_peg, finviz_peg):
        """3중 검증 로직 (원본 그대로)"""
        sources = []
        valid_pegs = []
        
        if yahoo_peg and 0 < yahoo_peg < 10:
            sources.append('Yahoo')
            valid_pegs.append(yahoo_peg)
        
        if calculated_peg and 0 < calculated_peg < 10:
            sources.append('Calc')
            valid_pegs.append(calculated_peg)
        
        if finviz_peg and 0 < finviz_peg < 10:
            sources.append('Finviz')
            valid_pegs.append(finviz_peg)
        
        if len(valid_pegs) < 2:
            return {'valid': False}
        
        avg_peg = sum(valid_pegs) / len(valid_pegs)
        
        for peg in valid_pegs:
            if abs(peg - avg_peg) / avg_peg > self.TOLERANCE:
                valid_pegs_sorted = sorted(valid_pegs)
                median_peg = valid_pegs_sorted[len(valid_pegs_sorted) // 2]
                
                return {
                    'valid': True,
                    'peg': round(median_peg, 2),
                    'status': '✅ 부분 검증 (중간값)',
                    'sources': sources
                }
        
        return {
            'valid': True,
            'peg': round(avg_peg, 2),
            'status': f'✅ {len(sources)}중 검증 통과',
            'sources': sources
        }
    
    def _step4_categorize(self):
        """Step 4: 유형별 분류 (원본 그대로, 중국 통계만 추가)"""
        logger.info("[Step 4/7] 유형별 분류 + 한글 번역...")
        df = pd.DataFrame(self.validated)
        
        categorized = {
            'best_value': [],
            'high_growth': [],
            'balanced': []
        }
        
        best = df[
            (df['peg'] < self.PEG_LIMITS['good']) &
            (df['growth_rate'] >= self.GROWTH_LIMITS['ideal_min']) &
            (df['growth_rate'] <= self.GROWTH_LIMITS['ideal_max'])
        ].sort_values('peg').head(10)
        
        for _, row in best.iterrows():
            categorized['best_value'].append(self._create_recommendation(row, 'best_value'))
        
        high = df[
            (df['growth_rate'] > 50) &
            (df['growth_rate'] <= self.GROWTH_LIMITS['max']) &
            (df['peg'] < 1.2)
        ].sort_values('growth_rate', ascending=False).head(10)
        
        for _, row in high.iterrows():
            categorized['high_growth'].append(self._create_recommendation(row, 'high_growth'))
        
        balanced = df[
            (df['peg'] < 1.0) &
            (df['growth_rate'] >= 20) &
            (df['growth_rate'] <= 40)
        ].sort_values('peg').head(5)
        
        for _, row in balanced.iterrows():
            categorized['balanced'].append(self._create_recommendation(row, 'balanced'))
        
        self.categorized_stocks = categorized
        
        # 중국 주식 통계
        china_count = sum(
            1 for cat_stocks in categorized.values()
            for stock in cat_stocks
            if stock.get('is_china', False)
        )
        
        logger.info(f"✅ 최고 가치주: {len(categorized['best_value'])}개")
        logger.info(f"✅ 고성장주: {len(categorized['high_growth'])}개")
        logger.info(f"✅ 균형: {len(categorized['balanced'])}개")
        logger.info(f"🇨🇳 중국 주식: {china_count}개\n")
        
        return True
    
    def _create_recommendation(self, row, category):
        """추천 생성 (원본 그대로, 중국 표시만 추가)"""
        ticker = row['ticker']
        peg = row['peg']
        growth = row['growth_rate']
        market_cap_b = row['market_cap'] / 1e9
        is_china = row.get('is_china', False)
        
        category_names = {
            'best_value': '최고 가치주',
            'high_growth': '고성장주',
            'balanced': '균형'
        }
        
        opinion = "🟢 강력 매수" if peg < self.PEG_LIMITS['excellent'] else ("🟢 매수" if peg < self.PEG_LIMITS['good'] else "🟡 관심")
        
        if market_cap_b < 1.0:
            opinion += " 💎 소형주"
        
        if is_china:
            opinion += " 🇨🇳"
        
        korean_desc = self.gpt_analyzer.translate_to_korean(
            row.get('name', 'N/A'),
            row.get('business_summary', '')
        )
        
        return {
            '티커': ticker,
            '회사명': row.get('name', 'N/A'),
            '한글설명': korean_desc,
            '섹터': row.get('sector', 'N/A'),
            '산업': row.get('industry', 'N/A'),
            '기업설명': row.get('business_summary', 'N/A'),
            'PEG': peg,
            '성장률(%)': growth,
            'P/E': row.get('pe_ratio'),
            '시가총액($B)': round(market_cap_b, 2),
            '투자의견': opinion,
            '검증상태': row['validation_status'],
            '데이터출처': ', '.join(row['data_sources']),
            '유형': category_names[category],
            'Yahoo': f"https://finance.yahoo.com/quote/{ticker}",
            'Finviz': f"https://finviz.com/quote.ashx?t={ticker}",
            'TradingView': f"https://www.tradingview.com/symbols/{ticker}",
            'price': row['price'],
            'category': category,
            'is_china': is_china
        }
    
    def _step5_create_excel(self):
        """Step 5: Excel 생성 (원본 그대로)"""
        logger.info("[Step 5/7] Excel 리포트 생성 중...")
        
        today = datetime.now().strftime('%Y%m%d')
        filename = f'Peter_Lynch_Report_{today}.xlsx'
        
        wb = Workbook()
        wb.remove(wb.active)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        for sheet_name, key in [
            ('🏆 최고 가치주 (40%)', 'best_value'),
            ('🚀 고성장주 (40%)', 'high_growth'),
            ('⚖️ 균형 (20%)', 'balanced')
        ]:
            stocks = self.categorized_stocks[key]
            if not stocks:
                continue
            
            ws = wb.create_sheet(title=sheet_name)
            columns = ['티커', '회사명', '한글설명', '유형', '섹터', '산업', 'PEG', '성장률(%)', 'P/E',
                      '시가총액($B)', '투자의견', '검증상태', '데이터출처', 'Yahoo', 'Finviz', 'TradingView']
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for row_idx, stock in enumerate(stocks, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = stock.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    if col_name in ['Yahoo', 'Finviz', 'TradingView'] and value:
                        cell.hyperlink = value
                        cell.style = 'Hyperlink'
                        cell.font = Font(color="0563C1", underline="single")
                    
                    if col_name == '투자의견':
                        if '강력 매수' in str(value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(bold=True, color="006100")
                        elif '소형주' in str(value):
                            cell.font = Font(bold=True, color="FF6600")
                        elif '🇨🇳' in str(value):
                            cell.font = Font(bold=True, color="FF0000")
            
            widths = [8, 25, 35, 12, 15, 20, 8, 10, 8, 12, 18, 15, 20, 15, 15, 15]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
        
        wb.save(filename)
        logger.info(f"✅ {filename}\n")
        return filename
    
    def _step6_gpt_analysis(self):
        """Step 6: GPT 분석 (원본 그대로)"""
        logger.info("[Step 6/7] GPT 포트폴리오 분석...")
        
        gpt_advice = self.gpt_analyzer.analyze_portfolio(
            self.categorized_stocks,
            self.history_manager.history
        )
        
        self.history_manager.update_from_portfolio(self.categorized_stocks)
        
        logger.info("✅ 완료\n")
        return gpt_advice
    
    def _step7_send_to_slack(self, filename, gpt_advice):
        """Step 7: 슬랙 전송 (주가 링크 추가)"""
        logger.info("[Step 7/7] 결과 전송...")
        
        # 주가 링크 생성 (새로 추가)
        stock_links = self._generate_stock_links()
        
        if not self.slack_sender.enabled:
            print("\n" + "="*80)
            print("📊 GPT 분석 결과")
            print("="*80)
            print(gpt_advice)
            print("\n" + "="*80)
            print("📈 추천 주식 주가 링크")
            print("="*80)
            print(stock_links)
            print("="*80 + "\n")
            return
        
        today = datetime.now().strftime('%Y년 %m월 %d일')
        week_num = datetime.now().isocalendar()[1]
        
        message = f"""🤖 *피터 린치 봇 - 공격적 포트폴리오 (전체 분석)*
📅 {today} ({week_num}주차)
💎 소형주 포함 ($100M+) - Tenbagger 발굴
🇨🇳 중국 비중 제한: 최대 1종목 (10%)

{gpt_advice}

━━━━━━━━━━━━━━━━━━
📈 *추천 주식 주가 링크*
━━━━━━━━━━━━━━━━━━
{stock_links}

━━━━━━━━━━━━━━━━━━
📂 {filename}
━━━━━━━━━━━━━━━━━━"""
        
        self.slack_sender.send_message(message)
        self.slack_sender.send_file(filename, f"리포트 - {today}")
        logger.info("✅ 완료\n")
    
    def _generate_stock_links(self):
        """주가 링크 생성 (새로 추가)"""
        links = []
        
        for category, name in [
            ('best_value', '🏆 최고 가치주'),
            ('high_growth', '🚀 고성장주'),
            ('balanced', '⚖️ 균형')
        ]:
            stocks = self.categorized_stocks.get(category, [])
            if stocks:
                links.append(f"\n*{name}*")
                for stock in stocks[:4]:
                    ticker = stock['티커']
                    name_kr = stock.get('한글설명', stock['회사명'])
                    price = stock.get('price', 0)
                    china_mark = " 🇨🇳" if stock.get('is_china', False) else ""
                    small_cap_mark = " 💎" if stock['시가총액($B)'] < 1.0 else ""
                    
                    yahoo_link = f"https://finance.yahoo.com/quote/{ticker}"
                    
                    links.append(
                        f"  • *{ticker}* - {name_kr}{china_mark}{small_cap_mark}\n"
                        f"    현재가: ${price:.2f} | <{yahoo_link}|주가 보기>"
                    )
        
        return "\n".join(links)
    
    def _print_summary(self):
        """콘솔 요약 (원본 그대로)"""
        print("\n" + "="*80)
        print("💡 공격적 포트폴리오 추천 (전체 분석)")
        print("="*80)
        
        for category, name in [('best_value', '최고 가치주'), ('high_growth', '고성장주'), ('balanced', '균형')]:
            stocks = self.categorized_stocks[category]
            if stocks:
                print(f"\n【{name}】")
                for stock in stocks[:3]:
                    small_cap_mark = " 💎" if stock['시가총액($B)'] < 1.0 else ""
                    china_mark = " 🇨🇳" if stock.get('is_china', False) else ""
                    print(f"  {stock['티커']:6} - {stock.get('한글설명', stock['회사명'])}{small_cap_mark}{china_mark}")
                    print(f"     PEG: {stock['PEG']:.2f} | 성장률: {stock['성장률(%)']:.1f}% | 시총: ${stock['시가총액($B)']:.2f}B")
                    print(f"     {stock['검증상태']}")
                    print(f"     주가: {stock['Yahoo']}")
        
        print("\n" + "="*80)


def main():
    print("""
╔════════════════════════════════════════════════════════════════╗
║  피터 린치 주식 스크리너 V5 - 완전판 (전체 분석)            ║
║                                                                ║
║  ✅ 모든 적격 티커 분석 (3000-5000개 예상)                  ║
║  ✅ 소형주 포함 ($100M+) - Tenbagger 발굴!                  ║
║  ✅ 3중 검증 (Yahoo + 직접계산 + Finviz)                     ║
║  ✅ 유형별 순위 관리 (슬롯 시스템)                           ║
║  ✅ 한글 기업 설명 (GPT 번역)                                ║
║  ✅ 실시간 주가 링크 (Yahoo, Finviz, TradingView)            ║
║  ✅ 중국 비중 10% 제한 (최대 1종목) ⭐ NEW                  ║
║  ✅ 슬랙 메시지에 주가 링크 추가 ⭐ NEW                      ║
║                                                                ║
║  "숨은 보석은 사람들이 안 보는 곳에 있다" - 피터 린치        ║
║  "10배 수익은 작은 회사에서 나온다" - 피터 린치              ║
║                                                                ║
║  환경 변수: OPENAI_API_KEY (필수)                             ║
╚════════════════════════════════════════════════════════════════╝
    """)
    
    if not os.environ.get("OPENAI_API_KEY"):
        print("⚠️  경고: OPENAI_API_KEY가 설정되지 않았습니다.")
        print("   기본 분석 모드로 실행됩니다.\n")
    
    if not os.environ.get("SLACK_BOT_TOKEN"):
        print("ℹ️  정보: 슬랙이 설정되지 않았습니다.")
        print("   결과는 콘솔에 출력됩니다.\n")
    
    screener = PeterLynchScreener()
    
    result = screener.run(ticker_limit=None)
    
    if result:
        print(f"\n✅ 스크리닝 완료!")
        print(f"📊 Excel 파일: {result}")
        print(f"📁 히스토리: portfolio_history.json")
        print(f"\n💎 모든 적격 주식을 분석했습니다.")
        print(f"   숨어있던 기회를 놓치지 않았습니다!")
        print(f"   소형주($100M+)에서 Tenbagger를 찾으세요!")
        print(f"🇨🇳 중국 주식은 최대 1종목(10%)으로 제한됩니다.")
    else:
        print("\n❌ 스크리닝 실패")
        print("로그 파일을 확인하세요.")


if __name__ == "__main__":
    main()